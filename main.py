import pandas as pandas
import dataclasses as dc
import openpyxl as xl


@dc.dataclass(unsafe_hash=True)
class DB_2:
    type: str = ""
    grade: str = ""
    sigugunNo: str = ""
    sigugunNm: str = ""
    category1: str = ""
    category2: str = ""
    value: int = ""


@dc.dataclass(unsafe_hash=True)
class DB_3:
    type: str = ""
    grade: str = ""
    sigugunNo: str = ""
    sigugunNm: str = ""
    store: str = ""
    value: int = ""


@dc.dataclass(unsafe_hash=True)
class DB_4:
    type: str = ""
    grade: str = ""
    sigugunNo: str = ""
    sigugunNm: str = ""
    region: str = ""


def getExecelDataBySheet(filePath, sheetName):
    df = pandas.read_excel(filePath, sheet_name=sheetName)

    data = []
    for idx, row in df.iterrows():

        if sheetName == "DB_2":
            if idx > 3 & idx < 256:
                data.extend(getDb2Format(row))
        elif sheetName == "DB_3":
            if idx > 3 & idx < 256:
                data.extend(getDb3Format(row))
        elif sheetName == "DB_4":
            data.extend(getDb4Format(row))

    return data


def getDb2Format(row):
    result = []

    # 원거리-Platinum Class(HH)
    result.append(DB_2("원거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], '관광자원 관심도', "자연", row.iloc[2]))
    result.append(DB_2("원거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], '관광자원 관심도', "인문", row.iloc[3]))
    result.append(DB_2("원거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], '관광자원 관심도', "레포츠", row.iloc[4]))
    result.append(DB_2("원거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], '관광자원 관심도', "쇼핑", row.iloc[5]))
    result.append(DB_2("원거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], '관광자원 관심도', "음식", row.iloc[6]))
    result.append(DB_2("원거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], '관광자원 관심도', "숙박", row.iloc[7]))
    result.append(DB_2("원거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], '관광자원 관심도', "추천코드", row.iloc[8]))
    result.append(DB_2("원거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], '관광 소비지출', "쇼핑", row.iloc[9]))
    result.append(DB_2("원거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], '관광 소비지출', "숙박", row.iloc[10]))
    result.append(DB_2("원거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], '관광 소비지출', "식음료", row.iloc[11]))
    result.append(DB_2("원거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], '관광 소비지출', "운송", row.iloc[12]))
    result.append(DB_2("원거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], '관광 소비지출', "여행", row.iloc[13]))
    result.append(DB_2("원거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], '관광 소비지출', "레저스포츠", row.iloc[14]))

    # 원거리-High Loyalty (LH)
    result.append(DB_2("원거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], '관광자원 관심도', "자연", row.iloc[15]))
    result.append(DB_2("원거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], '관광자원 관심도', "인문", row.iloc[16]))
    result.append(DB_2("원거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], '관광자원 관심도', "레포츠", row.iloc[17]))
    result.append(DB_2("원거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], '관광자원 관심도', "쇼핑", row.iloc[18]))
    result.append(DB_2("원거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], '관광자원 관심도', "음식", row.iloc[19]))
    result.append(DB_2("원거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], '관광자원 관심도', "숙박", row.iloc[20]))
    result.append(DB_2("원거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], '관광자원 관심도', "추천코드", row.iloc[21]))
    result.append(DB_2("원거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], '관광 소비지출', "쇼핑", row.iloc[22]))
    result.append(DB_2("원거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], '관광 소비지출', "숙박", row.iloc[23]))
    result.append(DB_2("원거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], '관광 소비지출', "식음료", row.iloc[24]))
    result.append(DB_2("원거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], '관광 소비지출', "운송", row.iloc[25]))
    result.append(DB_2("원거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], '관광 소비지출', "여행", row.iloc[26]))
    result.append(DB_2("원거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], '관광 소비지출', "레저스포츠", row.iloc[27]))

    # 원거리-More Promotion (HL)
    result.append(DB_2("원거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], '관광자원 관심도', "자연", row.iloc[28]))
    result.append(DB_2("원거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], '관광자원 관심도', "인문", row.iloc[29]))
    result.append(DB_2("원거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], '관광자원 관심도', "레포츠", row.iloc[30]))
    result.append(DB_2("원거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], '관광자원 관심도', "쇼핑", row.iloc[31]))
    result.append(DB_2("원거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], '관광자원 관심도', "음식", row.iloc[32]))
    result.append(DB_2("원거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], '관광자원 관심도', "숙박", row.iloc[33]))
    result.append(DB_2("원거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], '관광자원 관심도', "추천코드", row.iloc[34]))
    result.append(DB_2("원거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], '관광 소비지출', "쇼핑", row.iloc[35]))
    result.append(DB_2("원거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], '관광 소비지출', "숙박", row.iloc[36]))
    result.append(DB_2("원거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], '관광 소비지출', "식음료", row.iloc[37]))
    result.append(DB_2("원거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], '관광 소비지출', "운송", row.iloc[38]))
    result.append(DB_2("원거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], '관광 소비지출', "여행", row.iloc[29]))
    result.append(DB_2("원거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], '관광 소비지출', "레저스포츠", row.iloc[40]))

    # 원거리-Invest-Improve (LL)
    result.append(DB_2("원거리", "Invest-Improve (LL)", row.iloc[0], row.iloc[1], '관광자원 관심도', "자연", row.iloc[41]))
    result.append(DB_2("원거리", "Invest-Improve (LL)", row.iloc[0], row.iloc[1], '관광자원 관심도', "인문", row.iloc[42]))
    result.append(DB_2("원거리", "Invest-Improve (LL)", row.iloc[0], row.iloc[1], '관광자원 관심도', "레포츠", row.iloc[43]))
    result.append(DB_2("원거리", "Invest-Improve (LL)", row.iloc[0], row.iloc[1], '관광자원 관심도', "쇼핑", row.iloc[44]))
    result.append(DB_2("원거리", "Invest-Improve (LL)", row.iloc[0], row.iloc[1], '관광자원 관심도', "음식", row.iloc[45]))
    result.append(DB_2("원거리", "Invest-Improve (LL)", row.iloc[0], row.iloc[1], '관광자원 관심도', "숙박", row.iloc[46]))
    result.append(DB_2("원거리", "Invest-Improve (LL)", row.iloc[0], row.iloc[1], '관광자원 관심도', "추천코드", row.iloc[47]))
    result.append(DB_2("원거리", "Invest-Improve (LL)", row.iloc[0], row.iloc[1], '관광 소비지출', "쇼핑", row.iloc[48]))
    result.append(DB_2("원거리", "Invest-Improve (LL)", row.iloc[0], row.iloc[1], '관광 소비지출', "숙박", row.iloc[49]))
    result.append(DB_2("원거리", "Invest-Improve (LL)", row.iloc[0], row.iloc[1], '관광 소비지출', "식음료", row.iloc[50]))
    result.append(DB_2("원거리", "Invest-Improve (LL)", row.iloc[0], row.iloc[1], '관광 소비지출', "운송", row.iloc[51]))
    result.append(DB_2("원거리", "Invest-Improve (LL)", row.iloc[0], row.iloc[1], '관광 소비지출', "여행", row.iloc[52]))
    result.append(DB_2("원거리", "Invest-Improve (LL)", row.iloc[0], row.iloc[1], '관광 소비지출', "레저스포츠", row.iloc[53]))
    # ------------------------------------------------------------------------------------------------------------

    # 근거리-Platinum Class(HH)
    result.append(DB_2("근거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], '관광자원 관심도', "자연", row.iloc[54]))
    result.append(DB_2("근거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], '관광자원 관심도', "인문", row.iloc[55]))
    result.append(DB_2("근거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], '관광자원 관심도', "레포츠", row.iloc[56]))
    result.append(DB_2("근거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], '관광자원 관심도', "쇼핑", row.iloc[57]))
    result.append(DB_2("근거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], '관광자원 관심도', "음식", row.iloc[58]))
    result.append(DB_2("근거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], '관광자원 관심도', "숙박", row.iloc[59]))
    result.append(DB_2("근거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], '관광자원 관심도', "추천코드", row.iloc[60]))
    result.append(DB_2("근거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], '관광 소비지출', "쇼핑", row.iloc[61]))
    result.append(DB_2("근거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], '관광 소비지출', "숙박", row.iloc[62]))
    result.append(DB_2("근거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], '관광 소비지출', "식음료", row.iloc[63]))
    result.append(DB_2("근거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], '관광 소비지출', "운송", row.iloc[64]))
    result.append(DB_2("근거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], '관광 소비지출', "여행", row.iloc[65]))
    result.append(DB_2("근거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], '관광 소비지출', "레저스포츠", row.iloc[66]))

    # 근거리-High Loyalty (LH)
    result.append(DB_2("근거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], '관광자원 관심도', "자연", row.iloc[67]))
    result.append(DB_2("근거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], '관광자원 관심도', "인문", row.iloc[68]))
    result.append(DB_2("근거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], '관광자원 관심도', "레포츠", row.iloc[69]))
    result.append(DB_2("근거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], '관광자원 관심도', "쇼핑", row.iloc[70]))
    result.append(DB_2("근거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], '관광자원 관심도', "음식", row.iloc[71]))
    result.append(DB_2("근거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], '관광자원 관심도', "숙박", row.iloc[72]))
    result.append(DB_2("근거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], '관광자원 관심도', "추천코드", row.iloc[73]))
    result.append(DB_2("근거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], '관광 소비지출', "쇼핑", row.iloc[74]))
    result.append(DB_2("근거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], '관광 소비지출', "숙박", row.iloc[75]))
    result.append(DB_2("근거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], '관광 소비지출', "식음료", row.iloc[76]))
    result.append(DB_2("근거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], '관광 소비지출', "운송", row.iloc[77]))
    result.append(DB_2("근거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], '관광 소비지출', "여행", row.iloc[78]))
    result.append(DB_2("근거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], '관광 소비지출', "레저스포츠", row.iloc[79]))

    # 근거리-More Promotion (HL)
    result.append(DB_2("근거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], '관광자원 관심도', "자연", row.iloc[80]))
    result.append(DB_2("근거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], '관광자원 관심도', "인문", row.iloc[81]))
    result.append(DB_2("근거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], '관광자원 관심도', "레포츠", row.iloc[82]))
    result.append(DB_2("근거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], '관광자원 관심도', "쇼핑", row.iloc[81]))
    result.append(DB_2("근거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], '관광자원 관심도', "음식", row.iloc[84]))
    result.append(DB_2("근거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], '관광자원 관심도', "숙박", row.iloc[85]))
    result.append(DB_2("근거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], '관광자원 관심도', "추천코드", row.iloc[86]))
    result.append(DB_2("근거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], '관광 소비지출', "쇼핑", row.iloc[87]))
    result.append(DB_2("근거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], '관광 소비지출', "숙박", row.iloc[88]))
    result.append(DB_2("근거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], '관광 소비지출', "식음료", row.iloc[89]))
    result.append(DB_2("근거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], '관광 소비지출', "운송", row.iloc[90]))
    result.append(DB_2("근거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], '관광 소비지출', "여행", row.iloc[91]))
    result.append(DB_2("근거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], '관광 소비지출', "레저스포츠", row.iloc[92]))

    # 근거리-Invest-Improve (LL)
    result.append(DB_2("근거리", "Invest-Improve (LL)", row.iloc[0], row.iloc[1], '관광자원 관심도', "자연", row.iloc[93]))
    result.append(DB_2("근거리", "Invest-Improve (LL)", row.iloc[0], row.iloc[1], '관광자원 관심도', "인문", row.iloc[94]))
    result.append(DB_2("근거리", "Invest-Improve (LL)", row.iloc[0], row.iloc[1], '관광자원 관심도', "레포츠", row.iloc[95]))
    result.append(DB_2("근거리", "Invest-Improve (LL)", row.iloc[0], row.iloc[1], '관광자원 관심도', "쇼핑", row.iloc[96]))
    result.append(DB_2("근거리", "Invest-Improve (LL)", row.iloc[0], row.iloc[1], '관광자원 관심도', "음식", row.iloc[97]))
    result.append(DB_2("근거리", "Invest-Improve (LL)", row.iloc[0], row.iloc[1], '관광자원 관심도', "숙박", row.iloc[98]))
    result.append(DB_2("근거리", "Invest-Improve (LL)", row.iloc[0], row.iloc[1], '관광자원 관심도', "추천코드", row.iloc[99]))
    result.append(DB_2("근거리", "Invest-Improve (LL)", row.iloc[0], row.iloc[1], '관광 소비지출', "쇼핑", row.iloc[100]))
    result.append(DB_2("근거리", "Invest-Improve (LL)", row.iloc[0], row.iloc[1], '관광 소비지출', "숙박", row.iloc[101]))
    result.append(DB_2("근거리", "Invest-Improve (LL)", row.iloc[0], row.iloc[1], '관광 소비지출', "식음료", row.iloc[102]))
    result.append(DB_2("근거리", "Invest-Improve (LL)", row.iloc[0], row.iloc[1], '관광 소비지출', "운송", row.iloc[103]))
    result.append(DB_2("근거리", "Invest-Improve (LL)", row.iloc[0], row.iloc[1], '관광 소비지출', "여행", row.iloc[104]))
    result.append(DB_2("근거리", "Invest-Improve (LL)", row.iloc[0], row.iloc[1], '관광 소비지출', "레저스포츠", row.iloc[105]))
    # ------------------------------------------------------------------------------------------------------------

    return result


def getDb3Format(row):
    result = []

    # 원거리-Platinum Class(HH)
    result.append(DB_3("원거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], row.iloc[2], row.iloc[3]))
    result.append(DB_3("원거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], row.iloc[4], row.iloc[5]))
    result.append(DB_3("원거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], row.iloc[6], row.iloc[7]))
    result.append(DB_3("원거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], row.iloc[8], row.iloc[9]))
    result.append(DB_3("원거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], row.iloc[10], row.iloc[11]))
    result.append(DB_3("원거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], row.iloc[12], row.iloc[13]))
    result.append(DB_3("원거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], row.iloc[14], row.iloc[15]))
    result.append(DB_3("원거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], row.iloc[16], row.iloc[17]))
    result.append(DB_3("원거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], row.iloc[18], row.iloc[19]))
    result.append(DB_3("원거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], row.iloc[20], row.iloc[21]))

    # 원거리-High Loyalty (LH)
    result.append(DB_3("원거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], row.iloc[22], row.iloc[23]))
    result.append(DB_3("원거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], row.iloc[24], row.iloc[25]))
    result.append(DB_3("원거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], row.iloc[26], row.iloc[27]))
    result.append(DB_3("원거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], row.iloc[28], row.iloc[29]))
    result.append(DB_3("원거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], row.iloc[30], row.iloc[31]))
    result.append(DB_3("원거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], row.iloc[32], row.iloc[33]))
    result.append(DB_3("원거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], row.iloc[34], row.iloc[35]))
    result.append(DB_3("원거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], row.iloc[36], row.iloc[37]))
    result.append(DB_3("원거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], row.iloc[38], row.iloc[39]))
    result.append(DB_3("원거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], row.iloc[40], row.iloc[41]))

    # 원거리-High Loyalty (LH)
    result.append(DB_3("원거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], row.iloc[42], row.iloc[43]))
    result.append(DB_3("원거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], row.iloc[44], row.iloc[45]))
    result.append(DB_3("원거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], row.iloc[46], row.iloc[47]))
    result.append(DB_3("원거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], row.iloc[48], row.iloc[49]))
    result.append(DB_3("원거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], row.iloc[50], row.iloc[51]))
    result.append(DB_3("원거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], row.iloc[52], row.iloc[53]))
    result.append(DB_3("원거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], row.iloc[54], row.iloc[55]))
    result.append(DB_3("원거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], row.iloc[56], row.iloc[57]))
    result.append(DB_3("원거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], row.iloc[58], row.iloc[59]))
    result.append(DB_3("원거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], row.iloc[60], row.iloc[61]))

    # 원거리-High Loyalty (LH)
    result.append(DB_3("원거리", "Invest-Improve (LL) ", row.iloc[0], row.iloc[1], row.iloc[62], row.iloc[63]))
    result.append(DB_3("원거리", "Invest-Improve (LL) ", row.iloc[0], row.iloc[1], row.iloc[64], row.iloc[65]))
    result.append(DB_3("원거리", "Invest-Improve (LL) ", row.iloc[0], row.iloc[1], row.iloc[66], row.iloc[67]))
    result.append(DB_3("원거리", "Invest-Improve (LL) ", row.iloc[0], row.iloc[1], row.iloc[68], row.iloc[69]))
    result.append(DB_3("원거리", "Invest-Improve (LL) ", row.iloc[0], row.iloc[1], row.iloc[70], row.iloc[71]))
    result.append(DB_3("원거리", "Invest-Improve (LL) ", row.iloc[0], row.iloc[1], row.iloc[72], row.iloc[73]))
    result.append(DB_3("원거리", "Invest-Improve (LL) ", row.iloc[0], row.iloc[1], row.iloc[74], row.iloc[75]))
    result.append(DB_3("원거리", "Invest-Improve (LL) ", row.iloc[0], row.iloc[1], row.iloc[76], row.iloc[77]))
    result.append(DB_3("원거리", "Invest-Improve (LL) ", row.iloc[0], row.iloc[1], row.iloc[78], row.iloc[79]))
    result.append(DB_3("원거리", "Invest-Improve (LL) ", row.iloc[0], row.iloc[1], row.iloc[80], row.iloc[81]))
    # --------------------------------------------------------------------------------------------------------

    # 근거리-Platinum Class(HH)
    result.append(DB_3("근거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], row.iloc[82], row.iloc[83]))
    result.append(DB_3("근거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], row.iloc[84], row.iloc[85]))
    result.append(DB_3("근거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], row.iloc[86], row.iloc[87]))
    result.append(DB_3("근거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], row.iloc[88], row.iloc[89]))
    result.append(DB_3("근거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], row.iloc[90], row.iloc[91]))
    result.append(DB_3("근거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], row.iloc[92], row.iloc[93]))
    result.append(DB_3("근거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], row.iloc[94], row.iloc[95]))
    result.append(DB_3("근거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], row.iloc[96], row.iloc[97]))
    result.append(DB_3("근거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], row.iloc[98], row.iloc[99]))
    result.append(DB_3("근거리", "Platinum Class(HH)", row.iloc[0], row.iloc[1], row.iloc[100], row.iloc[101]))

    # 근거리-High Loyalty (LH)
    result.append(DB_3("근거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], row.iloc[102], row.iloc[103]))
    result.append(DB_3("근거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], row.iloc[104], row.iloc[105]))
    result.append(DB_3("근거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], row.iloc[106], row.iloc[107]))
    result.append(DB_3("근거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], row.iloc[108], row.iloc[109]))
    result.append(DB_3("근거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], row.iloc[110], row.iloc[111]))
    result.append(DB_3("근거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], row.iloc[112], row.iloc[113]))
    result.append(DB_3("근거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], row.iloc[114], row.iloc[115]))
    result.append(DB_3("근거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], row.iloc[116], row.iloc[117]))
    result.append(DB_3("근거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], row.iloc[118], row.iloc[119]))
    result.append(DB_3("근거리", "High Loyalty (LH)", row.iloc[0], row.iloc[1], row.iloc[120], row.iloc[121]))

    # 근거리-High Loyalty (LH)
    result.append(DB_3("근거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], row.iloc[122], row.iloc[123]))
    result.append(DB_3("근거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], row.iloc[124], row.iloc[125]))
    result.append(DB_3("근거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], row.iloc[126], row.iloc[127]))
    result.append(DB_3("근거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], row.iloc[128], row.iloc[129]))
    result.append(DB_3("근거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], row.iloc[130], row.iloc[131]))
    result.append(DB_3("근거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], row.iloc[132], row.iloc[133]))
    result.append(DB_3("근거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], row.iloc[134], row.iloc[135]))
    result.append(DB_3("근거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], row.iloc[136], row.iloc[137]))
    result.append(DB_3("근거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], row.iloc[138], row.iloc[139]))
    result.append(DB_3("근거리", "More Promotion (HL)", row.iloc[0], row.iloc[1], row.iloc[140], row.iloc[141]))

    # 근거리-High Loyalty (LH)
    result.append(DB_3("근거리", "Invest-Improve (LL) ", row.iloc[0], row.iloc[1], row.iloc[142], row.iloc[143]))
    result.append(DB_3("근거리", "Invest-Improve (LL) ", row.iloc[0], row.iloc[1], row.iloc[144], row.iloc[145]))
    result.append(DB_3("근거리", "Invest-Improve (LL) ", row.iloc[0], row.iloc[1], row.iloc[146], row.iloc[147]))
    result.append(DB_3("근거리", "Invest-Improve (LL) ", row.iloc[0], row.iloc[1], row.iloc[148], row.iloc[149]))
    result.append(DB_3("근거리", "Invest-Improve (LL) ", row.iloc[0], row.iloc[1], row.iloc[150], row.iloc[151]))
    result.append(DB_3("근거리", "Invest-Improve (LL) ", row.iloc[0], row.iloc[1], row.iloc[152], row.iloc[153]))
    result.append(DB_3("근거리", "Invest-Improve (LL) ", row.iloc[0], row.iloc[1], row.iloc[154], row.iloc[155]))
    result.append(DB_3("근거리", "Invest-Improve (LL) ", row.iloc[0], row.iloc[1], row.iloc[156], row.iloc[157]))
    result.append(DB_3("근거리", "Invest-Improve (LL) ", row.iloc[0], row.iloc[1], row.iloc[158], row.iloc[159]))
    result.append(DB_3("근거리", "Invest-Improve (LL) ", row.iloc[0], row.iloc[1], row.iloc[160], row.iloc[161]))
    # --------------------------------------------------------------------------------------------------------

    return result


def getDb4Format(row):
    result = []

    for i, column in enumerate(row):
        if i > 3:
            if not pandas.isnull(column):
                result.append(DB_4(row.iloc[0], row.iloc[1], row.iloc[2], row.iloc[3], column))
            else:
                break

    return result


def writeExcelDataBySheet(filePath, sheetName, data):
    workBook = xl.load_workbook(filePath)
    workSheet = workBook[sheetName]

    print("시트 : " + sheetName + "데이터 변환 시작")
    for i, item in enumerate(data):

        if sheetName == "관심도,소비지출":
            excelRowIndex = i + 6
            workSheet.cell(row=excelRowIndex, column=1).value = item.type
            workSheet.cell(row=excelRowIndex, column=2).value = item.grade
            workSheet.cell(row=excelRowIndex, column=3).value = item.sigugunNo
            workSheet.cell(row=excelRowIndex, column=4).value = item.sigugunNm
            workSheet.cell(row=excelRowIndex, column=5).value = item.category1
            workSheet.cell(row=excelRowIndex, column=6).value = item.category2
            workSheet.cell(row=excelRowIndex, column=7).value = item.value

        elif sheetName == "관심지점":
            excelRowIndex = i + 6
            workSheet.cell(row=excelRowIndex, column=1).value = item.type
            workSheet.cell(row=excelRowIndex, column=2).value = item.grade
            workSheet.cell(row=excelRowIndex, column=3).value = item.sigugunNo
            workSheet.cell(row=excelRowIndex, column=4).value = item.sigugunNm
            workSheet.cell(row=excelRowIndex, column=5).value = item.store
            workSheet.cell(row=excelRowIndex, column=6).value = item.value

        elif sheetName == "관련지역":
            excelRowIndex = i + 6
            workSheet.cell(row=excelRowIndex, column=1).value = item.type
            workSheet.cell(row=excelRowIndex, column=2).value = item.grade
            workSheet.cell(row=excelRowIndex, column=3).value = item.sigugunNo
            workSheet.cell(row=excelRowIndex, column=4).value = item.sigugunNm
            workSheet.cell(row=excelRowIndex, column=5).value = item.region

    workBook.save(filePath)
    workBook.close
    print("시트 : " + sheetName + "데이터 변환 종료")


if __name__ == '__main__':
    data = getExecelDataBySheet('0.(시각화의뢰)DATASET(240611).xlsx', 'DB_2')
    writeExcelDataBySheet("거리 등급 데이터.xlsx", "관심도,소비지출", data)

    data = getExecelDataBySheet('0.(시각화의뢰)DATASET(240611).xlsx', 'DB_3')
    writeExcelDataBySheet("거리 등급 데이터.xlsx", "관심지점", data)

    data = getExecelDataBySheet('0.(시각화의뢰)DATASET(240611).xlsx', 'DB_4')
    writeExcelDataBySheet("거리 등급 데이터.xlsx", "관련지역", data)
